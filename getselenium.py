import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class ResultBec:
#   """
#   Arama sonuçlarını tutmak için kullanılacak bir sınıf.
#   """
  def __init__(self, corporate_name, document_number, status, zip_code):
    self.corporate_name = corporate_name
    self.document_number = document_number
    self.status = status
    self.zip_code = zip_code

# Florida eyaleti için arama URL'si
wSearchTerm = 10036;
url = "https://search.sunbiz.org/Inquiry/CorporationSearch/SearchResults?inquiryType=ZipCode&searchTerm=" + str(wSearchTerm)

# Selenium WebDriver nesnesi oluştur
driver = webdriver.Chrome()

# Sayfayı aç
driver.get(url)

# Sonuçları bir listede tut
result_list = []
wTooMuch = 0
wEOF = 0

try:
    while wEOF == 0:
        if wTooMuch>0:
            button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Next List')))
            button.click()
            
        wTooMuch = wTooMuch + 1
        if wTooMuch > 99000:
            wEOF = 1
        # Tablo öğesini bekle
        table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "search-results")))

        # Tüm satırları bul
        rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Başlık satırını atla

        for row in rows:
            # Verileri ayıkla
            data_cells = row.find_elements(By.TAG_NAME, "td")
            corporate_name = data_cells[0].find_element(By.TAG_NAME, "a").text  # Şirket adını bağlantıdan al
            document_number = data_cells[1].text
            status = data_cells[2].text
            zip_code = data_cells[3].text

            # Sonuç nesnesi oluştur ve listeye ekle
            result = ResultBec(corporate_name, document_number, status, zip_code)
            result_list.append(result)
           

finally:
    # Sonuç listesini kullanın...
    # print(result_list)
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(str(wSearchTerm))
    # Başlık satırını yaz
    worksheet.cell(row=1, column=1).value = "Şirket Adı"
    worksheet.cell(row=1, column=2).value = "Sicil Numarası"
    worksheet.cell(row=1, column=3).value = "Durum"
    worksheet.cell(row=1, column=4).value = "Posta Kodu"

    # Sonuçları satır satır yaz
    for i, result in enumerate(result_list, start=2):
        worksheet.cell(row=i, column=1).value = result.corporate_name
        worksheet.cell(row=i, column=2).value = result.document_number
        worksheet.cell(row=i, column=3).value = result.status
        worksheet.cell(row=i, column=4).value = result.zip_code
    workbook.save("ZipKod.xlsx")
    # Her durumda tarayıcıyı kapat
    driver.quit()

